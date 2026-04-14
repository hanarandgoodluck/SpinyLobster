"""
文档解析服务 - 支持多种格式的需求文档解析
"""
import os
import logging
from typing import List, Dict
from abc import ABC, abstractmethod

logger = logging.getLogger(__name__)


class DocumentParser(ABC):
    """文档解析器基类"""
    
    @abstractmethod
    def parse(self, file_path: str) -> List[str]:
        """解析文档，返回文本段落列表"""
        pass


class TextParser(DocumentParser):
    """纯文本文件解析器"""
    
    def parse(self, file_path: str) -> List[str]:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # 按空行分割段落
            paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
            logger.info(f"文本文件解析完成，共 {len(paragraphs)} 个段落")
            return paragraphs
        except Exception as e:
            logger.error(f"文本文件解析失败: {str(e)}", exc_info=True)
            raise


class MarkdownParser(DocumentParser):
    """Markdown 文件解析器"""
    
    def parse(self, file_path: str) -> List[str]:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            paragraphs = []
            current_paragraph = []
            
            for line in content.split('\n'):
                line = line.strip()
                # 跳过标题行，但保留内容
                if line.startswith('#'):
                    if current_paragraph:
                        paragraphs.append(' '.join(current_paragraph))
                        current_paragraph = []
                    # 将标题作为独立段落
                    paragraphs.append(line)
                elif line == '':
                    if current_paragraph:
                        paragraphs.append(' '.join(current_paragraph))
                        current_paragraph = []
                else:
                    current_paragraph.append(line)
            
            if current_paragraph:
                paragraphs.append(' '.join(current_paragraph))
            
            logger.info(f"Markdown 文件解析完成，共 {len(paragraphs)} 个段落")
            return paragraphs
        except Exception as e:
            logger.error(f"Markdown 文件解析失败: {str(e)}", exc_info=True)
            raise


class DocxParser(DocumentParser):
    """Word 文档解析器"""
    
    def parse(self, file_path: str) -> List[str]:
        try:
            # 验证文件是否存在
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
            
            # 验证文件大小
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                raise ValueError(f"文件大小为 0: {file_path}")
            
            logger.info(f"开始解析 Word 文档: {file_path}, 大小: {file_size} bytes")
            
            from docx import Document
            
            doc = Document(file_path)
            paragraphs = []
            
            for para in doc.paragraphs:
                text = para.text.strip()
                if text:
                    paragraphs.append(text)
            
            logger.info(f"Word 文档解析完成，共 {len(paragraphs)} 个段落")
            return paragraphs
        except FileNotFoundError as e:
            logger.error(f"文件不存在: {str(e)}")
            raise
        except ValueError as e:
            logger.error(f"文件无效: {str(e)}")
            raise
        except ImportError:
            logger.error("python-docx 未安装，请运行: pip install python-docx")
            raise
        except Exception as e:
            logger.error(f"Word 文档解析失败: {str(e)}", exc_info=True)
            raise


class PdfParser(DocumentParser):
    """PDF 文档解析器"""
    
    def parse(self, file_path: str) -> List[str]:
        try:
            # 尝试使用 PyPDF2
            try:
                return self._parse_with_pypdf2(file_path)
            except ImportError:
                logger.warning("PyPDF2 未安装，尝试使用 pdfplumber")
                return self._parse_with_pdfplumber(file_path)
        except Exception as e:
            logger.error(f"PDF 文档解析失败: {str(e)}", exc_info=True)
            raise
    
    def _parse_with_pypdf2(self, file_path: str) -> List[str]:
        """使用 PyPDF2 解析 PDF"""
        from PyPDF2 import PdfReader
        
        reader = PdfReader(file_path)
        paragraphs = []
        
        for page in reader.pages:
            text = page.extract_text()
            if text:
                # 按行分割并过滤空行
                lines = [line.strip() for line in text.split('\n') if line.strip()]
                paragraphs.extend(lines)
        
        logger.info(f"PDF 文档解析完成（PyPDF2），共 {len(paragraphs)} 个段落")
        return paragraphs
    
    def _parse_with_pdfplumber(self, file_path: str) -> List[str]:
        """使用 pdfplumber 解析 PDF"""
        import pdfplumber
        
        paragraphs = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines = [line.strip() for line in text.split('\n') if line.strip()]
                    paragraphs.extend(lines)
        
        logger.info(f"PDF 文档解析完成（pdfplumber），共 {len(paragraphs)} 个段落")
        return paragraphs


class XlsxParser(DocumentParser):
    """Excel 文件解析器"""
    
    def parse(self, file_path: str) -> List[str]:
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True)
            paragraphs = []
            
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    # 合并非空单元格
                    cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if cells:
                        paragraphs.append(' | '.join(cells))
            
            wb.close()
            logger.info(f"Excel 文件解析完成，共 {len(paragraphs)} 个段落")
            return paragraphs
        except ImportError:
            logger.error("openpyxl 未安装，请运行: pip install openpyxl")
            raise
        except Exception as e:
            logger.error(f"Excel 文件解析失败: {str(e)}", exc_info=True)
            raise


class DocumentParserFactory:
    """文档解析器工厂"""
    
    PARSERS = {
        '.txt': TextParser,
        '.md': MarkdownParser,
        '.docx': DocxParser,
        '.pdf': PdfParser,
        '.xlsx': XlsxParser,
    }
    
    @classmethod
    def get_parser(cls, file_path: str) -> DocumentParser:
        """根据文件扩展名获取对应的解析器"""
        ext = os.path.splitext(file_path)[1].lower()
        parser_class = cls.PARSERS.get(ext)
        
        if not parser_class:
            raise ValueError(f"不支持的文件格式: {ext}")
        
        return parser_class()
    
    @classmethod
    def parse_file(cls, file_path: str) -> List[str]:
        """解析文件，返回文本段落列表"""
        parser = cls.get_parser(file_path)
        paragraphs = parser.parse(file_path)
        if not paragraphs:
            logger.warning(f"文件 {file_path} 解析结果为空，请检查文件内容")
        return paragraphs
